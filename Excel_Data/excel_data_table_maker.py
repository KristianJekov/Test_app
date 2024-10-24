import os
from colorama import Fore
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, PatternFill, Side


def get_bigger_version(version1, version2):
    # Replace second dot with hyphen
    v1 = version1.replace(".", "-", 2).replace("-", ".", 1)
    v2 = version2.replace(".", "-", 2).replace("-", ".", 1)

    # Convert to numeric parts for comparison
    v1_parts = [int(x) for x in v1.replace("-", ".").split(".")]
    v2_parts = [int(x) for x in v2.replace("-", ".").split(".")]

    # Compare each part
    for i in range(len(v1_parts)):
        if v1_parts[i] > v2_parts[i]:
            return v1
        elif v1_parts[i] < v2_parts[i]:
            return v2

    return v1


def create_file_for_data_table(amount, ver1, ver2):
    bigger_ver = get_bigger_version(ver1, ver2)

    # Assume the file directory is defined
    file_directory = r"D:\Documents\Custom Office Templates\Desktop\Update Tables Data"
    file_path = os.path.join(file_directory, f"Test Updates for v{bigger_ver}.xlsx")

    if not os.path.exists(file_directory):
        os.makedirs(file_directory)

    update_components(file_path, amount, ver1, ver2)


# Function to add the constant header lines at the top of every worksheet
def add_constant_header(ws):

    green_fill = PatternFill(
        start_color="26DE22", end_color="26DE22", fill_type="solid"
    )
    yellow_fill = PatternFill(
        start_color="F7EC09", end_color="F7EC09", fill_type="solid"
    )
    red_fill = PatternFill(start_color="EF4B11", end_color="EF4B11", fill_type="solid")

    ws["B2"].fill = green_fill
    ws["C2"] = "Test Passed"
    ws["B3"].fill = yellow_fill
    ws["C3"] = "Test Passed With Some Trouble"
    ws["B4"].fill = red_fill
    ws["C4"] = "Test Failed"


def edit_cell_size(ws):
    ws.column_dimensions["A"].width = 6
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 20
    ws.column_dimensions["G"].width = 50

    # Set initial row heights for the worksheet
    for row in range(1, 51):
        ws.row_dimensions[row].height = 20


def add_update_tables(ws, start_row, tests, ver1, ver2):
    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )

    def add_table(ws, start_row, ver1, ver2):
        header = [
            f"{ver1}->{ver2}",
            "Board",
            "Sensor Hub",
            "Battery",
            "Remote",
            "Notes",
        ]

        # Add the header row for the table
        for col, header_text in enumerate(header, start=2):  # Columns B to G
            ws.cell(row=start_row, column=col).value = header_text
            ws.cell(row=start_row, column=col).border = thick_border

        # Add test rows based on provided data
        for row_num in range(tests):
            current_row = start_row + 1 + row_num

            # Set "Test X" in column B (version field)
            ws.cell(row=current_row, column=2).value = f"Test {row_num + 1}"

            # Apply borders to all cells in the row (columns B through G)
            for col_idx in range(2, 8):  # 2 to 7 corresponds to columns B to G
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thick_border
                # Other cells remain empty but have borders

    # Add the first table (ver1 -> ver2)
    add_table(ws, start_row, ver1, ver2)

    # Add the second table (ver2 -> ver1)
    second_table_row = start_row + tests + 2  # Start second table after a gap
    add_table(ws, second_table_row, ver2, ver1)


# Function to update components, ensuring no overwriting of existing content
def update_components(
    file_path,
    amount,
    ver1,
    ver2,
):
    bigger_ver = get_bigger_version(ver1, ver2)

    if os.path.exists(file_path):
        # Load the existing workbook
        wb = load_workbook(file_path)
        ws = wb.active  # Assuming updates go to the active sheet
        print(Fore.CYAN + f"Workbook '{file_path}' found. Updating existing data.")
    else:
        # Create a new workbook and worksheet if it doesn't exist
        wb = Workbook()
        ws = wb.active
        ws.title = f"Updates on v{bigger_ver}"
        print(Fore.CYAN + f"New workbook created for '{file_path}'.")
        add_constant_header(ws)
        edit_cell_size(ws)  # Add header if new workbook

    # Find the next available row to avoid overwriting existing content
    start_row = ws.max_row + 2  # Leave one row gap for spacing
    add_update_tables(ws, start_row, amount, ver1, ver2)
    # Add the update table starting from the next available row
    # add_update_tables(ws, start_row, amount, ver1, ver2, progress_list)

    # Save the updated workbook
    wb.save(file_path)
    print(f"Excel file updated successfully at {file_path}")


def create_empty_worksheet(
    amount,
    ver1,
    ver2,
):

    create_file_for_data_table(amount, ver1, ver2)
    # add_update_tables(ws, start_row, tests, ver1, ver2, progress_list)


# Example usage


# # Progress list to be inserted into the table
# progress_list = ["1", "1", "1", "0", "N/A"]
# update_components(file_path, 5, "v2.2", "v2.3", progress_list)
